# Importing the openpyxl library to work with Excel files (.xlsx)
import openpyxl 
# Importing the tkinter library to create GUI (Graphical User Interface) applications
import tkinter as tk
# Importing ttk module from tkinter for themed widgets and messagebox module to display pop-up messages to the user
from tkinter import ttk, messagebox


# Function to load patient records from an Excel file
def load_records(filename="Patient_Data.xlsx"):
    
    try:
        wb = openpyxl.load_workbook(filename)  # Load the workbook
        sheet = wb.active  # Get the active sheet
        patients = []
        
        # Iterate through rows (starting from row 2 to skip headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            patients.append(row)  # Append each row to the patients list
        return patients
    
    except FileNotFoundError:  # If the file doesn't exist, return an empty list
        return []


# Function to save patient records to an Excel file
def save_records(patients, filename="Patient_Data.xlsx"):
    
    wb = openpyxl.Workbook()  # Create a new workbook
    sheet = wb.active  # Get the active sheet
    
    # Add headers to the sheet
    sheet.append(["Patient_ID", "Name", "Gender", "Room_Number", "Disease", "Age"])
    
    for patient in patients:  # Write patient records to the sheet
        sheet.append(patient)
    wb.save(filename)  # Save the workbook


# Function to add a new patient
def add_patient():
    
    def save_patient():
        # Retrieve data from input fields
        patient_id = patient_id_entry.get()
        name = name_entry.get()
        gender = gender_combobox.get()
        room_number = room_number_entry.get()
        disease = disease_entry.get()
        age = age_entry.get()

        # Check if all fields are filled
        if not all([patient_id, name, gender, room_number, disease, age]):
            messagebox.showerror("Error", "All fields are required!")
            return

        patients = load_records()  # Load existing records
        # Check if the patient ID already exists
        
        for patient in patients:
            if patient[0] == patient_id:
                messagebox.showerror("Error", "Patient ID already exists!")
                return

        # Add the new patient record
        new_patient = [patient_id, name, gender, int(room_number), disease, int(age)]
        patients.append(new_patient)
        save_records(patients)  # Save updated records
        messagebox.showinfo("Success", "Patient added successfully!")
        add_window.destroy()  # Close the add window

    # Create a new window for adding a patient
    add_window = tk.Toplevel(root)
    add_window.title("Add Patient")
    add_window.geometry("400x300")

    # Add input fields and labels
    tk.Label(add_window, text="Patient ID:").pack(pady=5)
    patient_id_entry = tk.Entry(add_window)
    patient_id_entry.pack()

    tk.Label(add_window, text="Name:").pack(pady=5)
    name_entry = tk.Entry(add_window)
    name_entry.pack()

    tk.Label(add_window, text="Gender:").pack(pady=5)
    gender_combobox = ttk.Combobox(add_window, values=["Male", "Female", "Other"])
    gender_combobox.pack()

    tk.Label(add_window, text="Room Number:").pack(pady=5)
    room_number_entry = tk.Entry(add_window)
    room_number_entry.pack()

    tk.Label(add_window, text="Disease:").pack(pady=5)
    disease_entry = tk.Entry(add_window)
    disease_entry.pack()

    tk.Label(add_window, text="Age:").pack(pady=5)
    age_entry = tk.Entry(add_window)
    age_entry.pack()

    # Add save button
    tk.Button(add_window, text="Save", command=save_patient).pack(pady=10)


# Function to view all patients
def view_patients():
    
    patients = load_records()  # Load patient records
    if not patients:  # If no records found, show info message
        messagebox.showinfo("Info", "No patient records found.")
        return

    # Create a new window for viewing patients
    view_window = tk.Toplevel(root)
    view_window.title("View Patients")
    view_window.geometry("600x400")

    # Create a treeview widget to display patient records
    tree = ttk.Treeview(view_window, columns=("ID", "Name", "Gender", "Room", "Disease", "Age"), show="headings")
    tree.heading("ID", text="Patient ID")
    tree.heading("Name", text="Name")
    tree.heading("Gender", text="Gender")
    tree.heading("Room", text="Room Number")
    tree.heading("Disease", text="Disease")
    tree.heading("Age", text="Age")
    tree.pack(fill=tk.BOTH, expand=True)

    # Populate the treeview with patient records
    for patient in patients:
        tree.insert("", tk.END, values=patient)


# Function to search for a specific patient
def search_patient():
    
    def search():
        search_id = search_entry.get()  # Get the patient ID to search
        if not search_id:  # If no ID entered, show error
            messagebox.showerror("Error", "Please enter a Patient ID.")
            return

        patients = load_records()  # Load patient records
        for patient in patients:  # Search for the patient
            if patient[0] == search_id:
                # Display patient details if found
                result = f"Patient_ID: {patient[0]}, Name: {patient[1]}, Gender: {patient[2]}, Room: {patient[3]}, Disease: {patient[4]}, Age: {patient[5]}"
                messagebox.showinfo("Patient Found", result)
                return

        messagebox.showinfo("Not Found", "Patient not found.")  # If not found, show info

    # Create a new window for searching a patient
    search_window = tk.Toplevel(root)
    search_window.title("Search Patient")
    search_window.geometry("300x200")

    tk.Label(search_window, text="Enter Patient ID:").pack(pady=5)
    search_entry = tk.Entry(search_window)
    search_entry.pack()

    tk.Button(search_window, text="Search", command=search).pack(pady=10)


# Function to delete a patient record
def delete_patient():
    
    def delete():
        delete_id = delete_entry.get()  # Get the patient ID to delete
        if not delete_id:  # If no ID entered, show error
            messagebox.showerror("Error", "Please enter a Patient ID.")
            return

        patients = load_records()  # Load patient records
        # Filter out the patient to be deleted
        updated_patients = [patient for patient in patients if patient[0] != delete_id]

        if len(updated_patients) == len(patients):  # If no record was deleted
            messagebox.showinfo("Not Found", "Patient not found.")
        else:
            save_records(updated_patients)  # Save updated records
            messagebox.showinfo("Success", "Patient deleted successfully!")
            delete_window.destroy()  # Close the delete window

    # Create a new window for deleting a patient
    delete_window = tk.Toplevel(root)
    delete_window.title("Delete Patient")
    delete_window.geometry("300x200")

    tk.Label(delete_window, text="Enter Patient ID:").pack(pady=5)
    delete_entry = tk.Entry(delete_window)
    delete_entry.pack()

    tk.Button(delete_window, text="Delete", command=delete).pack(pady=10)


# Function to update an existing patient record
def update_patient():
    
    def update():
        update_id = update_id_entry.get()  # Get the patient ID to update
        if not update_id:  # If no ID entered, show error
            messagebox.showerror("Error", "Please enter a Patient ID.")
            return

        patients = load_records()  # Load patient records
        found = False  # Flag to check if patient exists
        updated_patients = []

        for patient in patients:
            if patient[0] == update_id:  # If patient found
                found = True
                # Get updated values or retain old values
                name = name_entry.get() or patient[1]
                gender = gender_combobox.get() or patient[2]
                room_number = room_number_entry.get() or patient[3]
                disease = disease_entry.get() or patient[4]
                age = age_entry.get() or patient[5]
                # Append updated record
                updated_patients.append([patient[0], name, gender, int(room_number), disease, int(age)])
            else:
                updated_patients.append(patient)

        if not found:  # If patient not found
            messagebox.showinfo("Not Found", "Patient not found.")
        else:
            save_records(updated_patients)  # Save updated records
            messagebox.showinfo("Success", "Patient updated successfully!")
            update_window.destroy()  # Close the update window

    # Create a new window for updating a patient
    update_window = tk.Toplevel(root)
    update_window.title("Update Patient")
    update_window.geometry("400x300")

    # Add input fields and labels
    tk.Label(update_window, text="Enter Patient ID:").pack(pady=5)
    update_id_entry = tk.Entry(update_window)
    update_id_entry.pack()

    tk.Label(update_window, text="Name:").pack(pady=5)
    name_entry = tk.Entry(update_window)
    name_entry.pack()

    tk.Label(update_window, text="Gender:").pack(pady=5)
    gender_combobox = ttk.Combobox(update_window, values=["Male", "Female", "Other"])
    gender_combobox.pack()

    tk.Label(update_window, text="Room Number:").pack(pady=5)
    room_number_entry = tk.Entry(update_window)
    room_number_entry.pack()

    tk.Label(update_window, text="Disease:").pack(pady=5)
    disease_entry = tk.Entry(update_window)
    disease_entry.pack()

    tk.Label(update_window, text="Age:").pack(pady=5)
    age_entry = tk.Entry(update_window)
    age_entry.pack()

    # Add update button
    tk.Button(update_window, text="Update", command=update).pack(pady=10)

# Main Tkinter window
root = tk.Tk()
root.title("Hospital Management System")
root.geometry("400x400")

# Add main window widgets
tk.Label(root, text="Hospital Management System", font=("Arial", 16)).pack(pady=20)

tk.Button(root, text="Add Patient", command=add_patient).pack(pady=10)
tk.Button(root, text="View Patients(Any Name)", command=view_patients).pack(pady=10)
tk.Button(root, text="Search Patient", command=search_patient).pack(pady=10)
tk.Button(root, text="Update Patient", command=update_patient).pack(pady=10)
tk.Button(root, text="Delete Patient", command=delete_patient).pack(pady=10)

# Run the main event loop
root.mainloop()