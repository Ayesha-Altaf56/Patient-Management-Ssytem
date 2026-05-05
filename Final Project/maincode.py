import openpyxl  # Import the openpyxl library, which allows working with Excel files (.xlsx)
from openpyxl import Workbook  # Import the Workbook class from openpyxl to create a new Excel workbook

file_name = "Patient_Data.xlsx"  # Define the filename for patient records

# Function to load patient records from an Excel file
def load_records(Patient_Record):  
    
    try:
        # Load the workbook and select the active sheet
        wb = openpyxl.load_workbook(file_name)
        sheet=wb.active
        
        # Initialize an empty list to hold patient records
        patients=[] 
        
        # Loop through rows, starting from the second row (to skip the header) 
        for row in sheet.iter_rows(min_row=2, values_only=True):
            patients.append(row)  # Append each row (patient data) to the list
        
        return patients  # Return the list of patient records
    
    except FileNotFoundError:
        # If file is not found, return an empty list
        return []
        

# Main menu function to interact with the user       
def main_menue():  

    patients = load_records(file_name) # Load existing patient records
    
    while True:
        
        # Display menu options for the user
        print("\nHospital Management System")
        print("1. Add Patient")
        print("2. View All Patients")
        print("3. Search Patient by ID")
        print("4. Update Patient Info")
        print("5. Delete Patient")
        print("6. Save and Exit")
        
        # Get the user's choice
        choice=int(input("Enter the choice via their respective numbers: "))
        
        # Perform actions based on the user's choice
        if choice==1:
            add_patient(patients)  # Add new patient
        elif choice==2:
            view_patients(patients)  # View all patients
        elif choice==3:
            search_patient(patients)  # Search for a patient by ID
        elif choice==4:
            update_patient()  # Update a patient's information
        elif choice==5:
            delete_patient()  # Delete a patient record
        elif choice==6:
            save_records(file_name, patients)  # Save all records and exit
            print("Record saves. Exiting the system.")
        else: print("Invalid choice, please try again")
        
        break  # Exit the loop (and the program)
            

# Function to add a new patient to the records                
def add_patient(patients):  

    patients = load_records(file_name)  # Load existing patient records
    
    # Load the Excel workbook and access the active sheet
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    
    Patient_ID = input("Enter the patient ID: ")  # Get the patient's ID
    
    # Check if the patient ID already exists in the records
    for patient in patients:
        if patient[0]==Patient_ID:  
            print("Patient is already registered")  # Patient is already in the records
            return
    
    # Get other patient details
    Name = input("Enter the name of the patient: ")
    Gender = input("Enter the gender of the patient: ")
    Room_Number = int(input("Enter the room number of the patient: "))
    Disease = input("Enter the disease of the patient: ")
    Age = int(input("Enter the age of the patient: "))
    
    #  Append the new patient record to the sheet
    sheet.append([Patient_ID, Name, Gender, Room_Number, Disease, Age])
    print("Patient added successfully.")
    
    # Save the updated workbook
    wb.save(file_name)
 
 
# Function to view all patients in the records   
def view_patients(patients):

    patients = load_records(file_name)  # Load existing patient records
    
    wb = openpyxl.load_workbook(file_name)  # Load the Excel workbook
    sheet = wb.active
    
    print("Patient Records:")  # Display the heading
    
    # Display a message if there are no patients
    if not patients:
        print("No patient record found")
    else:
        print("\nPatient Record:")
        # Loop through all patients and print their details
        for patient in patients:
            print("Patient_ID:" ,patient[0], "Name:", patient[1], "Gender:", patient[2], "Room_Number:" ,patient[3], "Disease:" ,patient[4], "Age:" ,patient[5])


# Function to search for a patient by ID
def search_patient(patients):

    patients = load_records(file_name)  # Load existing patient records
    
    search_id=input("Enter patient ID to search: ")  # Get the patient ID to search for
    found=False
    
    # Loop through all patients to find the matching ID
    for patient in patients:
        
        if patient[0]==search_id:  # Print the patient's details if found
            print(f"Patient_ID : {patient[0]}, Name: {patient[1]}, Gender: {patient[2]}, Room_Number: {patient[3]}, Disease: {patient[4]}, Age: {patient[5]}")
            found=True
            
            break  # Exit the loop after finding the patient
        
    if not found:
        print("Patient not found")  # Display message if the patient is not found


# Function to update patient information
def update_patient():

    patients = load_records(file_name) # Load existing patient records
    
    search_id = input("Enter the patient ID to update: ")  # Get the patient ID to update
    found = False
    
    # Create a new workbook to save updated records
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Patient_ID", "Name", "Gender", "Room_Number", "Disease", "Age"])  # Header row

    # Loop through all patients to find the one to update
    for patient in patients:
        
        if search_id == patient[0]:
            found = True
            print(f"Current Info: Patient_ID: {patient[0]}, Name: {patient[1]}, Gender: {patient[2]}, Room_Number: {patient[3]}, Disease: {patient[4]}, Age: {patient[5]}")
            
            # Prompt for new information
            Name = input("Enter the new name (leave blank to keep current): ")
            Gender = input("Enter the new gender (leave blank to keep current): ")
            Room_Number = input("Enter the new room number (leave blank to keep current): ")
            Disease = input("Enter the new disease (leave blank to keep current): ")
            Age = input("Enter the new age (leave blank to keep current): ")
            
            # Update patient information, keeping current values if input is blank
            updated_patient = (
                patient[0],  # Patient_ID
                Name if Name else patient[1],  # Name
                Gender if Gender else patient[2],  # Gender
                Room_Number if Room_Number else patient[3],  # Room_Number
                Disease if Disease else patient[4],  # Disease
                Age if Age else patient[5]  # Age
            )
            sheet.append(updated_patient)  # Add updated patient to the new sheet
            print("Patient information updated!")
        else:
            sheet.append(patient)  # Add other patients to the new sheet

    if not found:
        print("Patient not found.")
    
    # Save the updated records back to the file
    wb.save(file_name)
    

# Function to delete a patient record    
def delete_patient():

    patients = load_records(file_name)  # Load existing patient records
    
    search_ID = input("Enter the patient's ID to remove: ")  # Get the patient ID to delete
    found = False
    
    # Create a new workbook to save updated records
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Patient_ID", "Name", "Gender", "Room_Number", "Disease", "Age"])  # # Add header row to the new sheet

    # Loop through all patients and remove the matching record
    for patient in patients:
        
        if search_ID == patient[0]:
            found = True
            print("Patient Record deleted.")
            continue  # Skip adding this patient to the new sheet
        sheet.append(patient)  # Add other patients to the new sheet

    if not found:
        print("Patient not found.")
    
    # Save the updated records back to the file
    wb.save(file_name)


# Function to save patient records to the Excel file      
def save_records(file_name, patients):
    
    # Load the existing workbook to update with the latest records
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    
    # Loop through all patient records and append them to the sheet
    for patient in patients:
        sheet.append(patient)
    
    # Save the workbook with the updated records
    wb.save(file_name)
    
        
# Example usage
main_menue()